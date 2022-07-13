import openpyxl
from openpyxl import Workbook
import openpyxl, pprint

libro = openpyxl.load_workbook(filename='Deudas.xlsx')

hoja = libro["Hoja1"]

clientes = {}

for fila in range(2, hoja.max_row + 1):
    cliente = hoja.cell(row=fila, column=3).value
    importe = hoja.cell(row=fila, column=4).value
    clientes.setdefault(cliente,0)
    clientes[cliente] += importe


for fila in range(2, hoja.max_row + 1):
    cliente = hoja.cell(row=fila, column=3).value
    importe = hoja.cell(row=fila, column=5)
    importe.value = clientes[cliente]

libro.save("deudas2.xlsx")


pprint.pp(clientes)














